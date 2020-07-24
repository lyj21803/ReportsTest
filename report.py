import openpyxl
import xlrd
import wx
import time
import os
import datetime
from datetime import date

#通过对话框获取路径
def get_path(wildcard, prompt):
    app = wx.App(None)
    style = wx.FD_OPEN | wx.FD_FILE_MUST_EXIST
    dialog = wx.FileDialog(None, prompt, wildcard=wildcard, style=style)
    if dialog.ShowModal() == wx.ID_OK:
        path = dialog.GetPath()
    else:
        path = None
    dialog.Destroy()
    return path

#获得工作簿中每一张工作表的最后一列数据地址...
def get_Excel_lastcols (excelfile):
    wb=xlrd.open_workbook(excelfile, 'r')
    sheetname=[]
    finalcol=[]
    for sheet in wb.sheets():
        if sheet.name=="沉降类":

            for col in range(0, sheet.ncols):
                if sheet.cell(0, col).ctype == 0:
                    finalcol.append(col-1)
                    break

        elif sheet.name=="水平位移类":
            for col in range(0, sheet.ncols,2):
                if sheet.cell(0, col).ctype == 0:
                    finalcol.append(col - 1)
                    break
        elif sheet.name[0:2]=="CX":
            finalcol.append(sheet.ncols-1)
        sheetname.append(sheet.name)

    return sheetname, finalcol

#获得工作表的最后一列数据地址...
def get_sheet_lastcol(sheet):
    #finalcol=0
    if sheet.title=="沉降类":
        for col in sheet.iter_rows(min_row=0,max_row=1,min_col=0,max_col=sheet.max_column,values_only=True):
            for i in range(0,len(col)):
                if col[i] == None:
                    return i-1
              
    elif sheet.title=="水平位移类":
        for col in sheet.iter_rows(min_row=0,max_row=1,min_col=0,max_col=sheet.max_column,values_only=True):
            for i in range(0,len(col),2):
                if col[i] == None:
                    return i-1
                 
    elif sheet.title[0:2]=="CX":
        for col in sheet.iter_rows(min_row=0,max_row=1,min_col=0,max_col=sheet.max_column,values_only=True):
            for i in range(0,len(col)):
                if col[i] == None:
                    return i-1
          
def get_datas_From_sheet(sheet,mincol,maxcol):
    datas=[]
    for col in  sheet.iter_cols(min_row=0,max_row=sheet.max_row,min_col=mincol,max_col=maxcol,values_only=True):
        datas.append(col)
    return datas    
#根据汇总数据写入报表，不同项目需要重新写入(沉降和水平位移类数据)
def write_datas_To_workbook(wb,report):
    workbook=openpyxl.load_workbook(report)
    #----------------------------------------------------------------------
    ws=wb["沉降类"]
    Highscol=get_sheet_lastcol(ws)
    if Highscol==None:
        Highscol=ws.max_column-1
    datas=get_datas_From_sheet(ws,Highscol,Highscol+1)
    worksheet=workbook["封面"]
    worksheet['B34']=(datas[1])[0]
    print("正在写入水位数据")
    worksheet=workbook["水位监测报表"]
    for row in range(9,13):
        for col in range(6,8):
            worksheet.cell(row,col,value=(datas[col-6])[row+60])
    print("正在写入坡顶沉降数据")
    worksheet=workbook["坡顶沉降监测报表"]
    for row in range(9,25):
        for col in range(5,7):
            worksheet.cell(row,col,value=(datas[col-5])[row+32])
    print("正在写入坡顶河坎沉降数据")
    worksheet=workbook["河坎沉降监测"]
    for row in range(9,13):
        for col in range(5,7):
            worksheet.cell(row,col,value=(datas[col-5])[row+56])
    print("正在写入祠堂沉降数据")
    worksheet=workbook["祠堂沉降监测"]
    for row in range(9,17):
        for col in range(5,7):
            worksheet.cell(row,col,value=(datas[col-5])[row+48])    
    print("正在写入周边地表沉降数据")
    worksheet=workbook["周边地表沉降监测"]
    for row in range(9,19):
        for col in range(5,7):
            worksheet.cell(row,col,value=(datas[col-5])[row-8])   
    for row in range(9,19):
        for col in range(12,14):
            worksheet.cell(row,col,value=(datas[col-12])[row+2])      
    for row in range(31,41):
        for col in range(5,7):
            worksheet.cell(row,col,value=(datas[col-5])[row-10])   
    for row in range(31,41):
        for col in range(12,14):
            worksheet.cell(row,col,value=(datas[col-12])[row])    
    print("沉降类数据写入报表完成！") 
    #------------------------------------------------------------------------------
    ws=wb["水平位移类"]
    Planecol=get_sheet_lastcol(ws)
    if Planecol==None:
        Planecol=ws.max_column-1
    planedatas=get_datas_From_sheet(ws,Planecol-2,Planecol+1)
    #print(planedatas)
    print("获得水平位移数据成功，准备写入报表...")
    worksheet=workbook["坡顶水平位移监测报表"]
    for row in range(8,24):
        for col in range(6,10):
            worksheet.cell(row,col,value=(planedatas[col-6])[row-6])  

    worksheet=workbook["河坎水平位移监测"]
    for row in range(8,12):
        for col in range(6,10):
            worksheet.cell(row,col,value=(planedatas[col-6])[row+10])
    #-----------------------------------------------------------------------
   
    workbook.save(filename=report)
#读入测斜数据    
def write_CXDatas_To_workbook(wb,report):
    workbook=openpyxl.load_workbook(report)
     #-----------------------------------------
    for sheet in wb:
        if sheet.title[0:2]=="CX":
            ws=wb[sheet.title]
            lastcol=get_sheet_lastcol(ws)
            if lastcol==None:
                lastcol=ws.max_column-1

            datas=get_datas_From_sheet(ws,lastcol,lastcol+1)
            temp=workbook["封面"]
            
            if (datas[1])[0]!=temp['B34'].value: 
               continue

            print("开始写入测斜数据："+sheet.title)
            worksheet=workbook[sheet.title]
            for row in range(10,10+len(datas[0])):
                for col in range(2,4):
                    worksheet.cell(row,col,value=(datas[col-2])[row-10])
            
      #-----------------------------------------------------------------------
    print("数据写入完成！")
    workbook.save(filename=report)      

        


#程序主接口...
if  __name__ == '__main__':
    excelDatas=get_path("*.xlsx","请选择数据汇总文件...")
    wb=openpyxl.load_workbook(excelDatas)

    report=get_path("*.xlsx","请选择报表模版文件...")
    write_datas_To_workbook(wb,report)

    write_CXDatas_To_workbook(wb,report)
    
   